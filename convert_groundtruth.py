"""
convert_groundtruth.py
----------------------
Converts a ground truth spreadsheet (CSV or Excel) to JSON format
matching the detection results output.

Usage:
    python convert_groundtruth.py --input ground_truth.xlsx --output ground_truth.json
    python convert_groundtruth.py --input ground_truth.csv --output ground_truth.json
"""

import argparse
import json
import sys
from pathlib import Path

def load_spreadsheet(path):
    """Load CSV or Excel file into a list of dicts."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        import csv
        with open(path, encoding="utf-8") as f:
            return list(csv.DictReader(f))
    elif suffix in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            sys.exit("openpyxl not installed. Run: pip install openpyxl")
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
        return rows
    else:
        sys.exit(f"Unsupported file format: {suffix}. Use .csv or .xlsx")

def parse_bool(value):
    """Parse TRUE/FALSE strings or boolean values."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().upper() in ("TRUE", "1", "YES")
    return False

def convert(rows):
    """Group rows by file and build ground truth structure."""
    files = {}
    document_notes = {}

    for row in rows:
        filename = str(row.get("file", "")).strip()
        text = str(row.get("text", "")).strip() if row.get("text") else ""

        if not filename:
            continue

        # Store document-level note once per file
        doc_note = str(row.get("document_notes", "")).strip() if row.get("document_notes") else ""
        if doc_note and filename not in document_notes:
            document_notes[filename] = doc_note

        if filename not in files:
            files[filename] = []

        # Skip rows with no text (letters noted as having no underlining)
        if not text:
            continue

        entry = {
            "text": text,
            "spans_multiple_words": parse_bool(row.get("spans_multiple_words")),
            "spans_line_break": parse_bool(row.get("spans_line_break")),
            "confidence": str(row.get("confidence", "certain")).strip().lower(),
            "notes": str(row.get("notes", "")).strip() if row.get("notes") else ""
        }
        files[filename].append(entry)

    # Build final structure
    ground_truth = []
    for filename, underlinings in files.items():
        entry = {
            "file": filename,
            "underlinings": underlinings
        }
        if filename in document_notes:
            entry["document_notes"] = document_notes[filename]
        ground_truth.append(entry)

    # Sort by filename
    ground_truth.sort(key=lambda x: x["file"])
    return {"ground_truth": ground_truth}

def main():
    p = argparse.ArgumentParser(description="Convert ground truth spreadsheet to JSON.")
    p.add_argument("--input", required=True, help="Input spreadsheet (.csv or .xlsx)")
    p.add_argument("--output", default="ground_truth.json", help="Output JSON file")
    args = p.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        sys.exit(f"File not found: {input_path}")

    rows = load_spreadsheet(input_path)
    print(f"Loaded {len(rows)} rows from {input_path.name}")

    result = convert(rows)

    total_files = len(result["ground_truth"])
    total_underlinings = sum(len(f["underlinings"]) for f in result["ground_truth"])
    print(f"Files: {total_files}")
    print(f"Total underlinings: {total_underlinings}")

    output_path = Path(args.output)
    output_path.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Saved to {output_path}")

if __name__ == "__main__":
    main()
